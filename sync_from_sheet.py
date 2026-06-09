"""
St Clair Cannabis - Auto Sync from Google Sheets
==========================================
Reads master product data (FLOWERS_LIVE, ITEMS_LIVE) and cross-references
with store-specific flags (FlowersFlags_POS, ItemsFlags_POS) to generate
a filtered menu for St Clair Cannabis (ALC01).

Two Google Sheets:
  1. MENU sheet  - contains FLOWERS_LIVE, ITEMS_LIVE (master product catalog)
  2. POS sheet   - contains FlowersFlags_POS, ItemsFlags_POS (store visibility)

Setup (one-time):
  1. Both sheets must be publicly readable (Share > Anyone with link > Viewer)
  2. Set MENU_SHEET_ID and POS_SHEET_ID below
  3. Run: python sync_from_sheet.py
  4. Or auto-deploy: python sync_from_sheet.py --deploy
"""

import csv
import io
import json
import os
import re
import subprocess
import sys
import urllib.request
import urllib.error

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FLOWERS_OUT = os.path.join(SCRIPT_DIR, "app", "lib", "flowers.json")
ITEMS_OUT = os.path.join(SCRIPT_DIR, "app", "lib", "items.json")

# ============================================================
#   CONFIGURATION - Set these values
# ============================================================
STORE_CODE = "STC01"  # Store code

# Google Sheet IDs (the long string in the URL)
# https://docs.google.com/spreadsheets/d/SHEET_ID_HERE/edit
MENU_SHEET_ID = "1ECyzLymF6-aZn30Lt_BTzvXvXZr6LlEPHVixiv3McbQ"   # SETB Anti v2.3 6G MENU (FLOWERS_LIVE, ITEMS_LIVE)
POS_SHEET_ID = ""    # POS sheet (FlowersFlags_POS, ItemsFlags_POS)
# ============================================================

# Sheet tab names
FLOWERS_SHEET = "FLOWERS_LIVE"
ITEMS_SHEET = "ITEMS_LIVE"
FLOWER_FLAGS_SHEET = "FlowersFlags_POS"
ITEM_FLAGS_SHEET = "ItemsFlags_POS"


def fetch_sheet_csv(sheet_id: str, sheet_name: str) -> list[dict]:
    """Fetch a Google Sheet tab as CSV and return list of dicts."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/gviz/tq?tqx=out:csv&sheet={urllib.request.quote(sheet_name)}"
    )
    print(f"  Fetching {sheet_name}...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            text = resp.read().decode("utf-8-sig")
    except urllib.error.URLError as e:
        print(f"  ERROR: Could not fetch '{sheet_name}': {e}")
        print(f"  Make sure the sheet is shared as 'Anyone with link'")
        return []

    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not any(v.strip() for v in row.values()):
            continue
        rows.append(row)
    return rows


def read_xlsx_sheet(wb, sheet_name: str) -> list[dict]:
    """Read a sheet from an openpyxl workbook as list of dicts."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found in XLSX")
        return []
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                val = row[i]
                d[h] = str(val) if val is not None else ""
            elif h:
                d[h] = ""
        rows.append(d)
    return rows


# ── Helpers ──

def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"[\s-]+", "-", s)
    return s.strip("-")


def parse_single_price(val: str):
    if not val or val.strip() in ("", "-", "N/A", "None"):
        return None
    try:
        return int(float(val.strip()))
    except ValueError:
        return None

def parse_price_point(val: str):
    if not val or val.strip() in ("", "-", "N/A", "None"):
        return None
    val_str = str(val).strip().replace("$", "").replace(",", "")
    if "|" in val_str:
        parts = val_str.split("|")
        reg_val = parse_single_price(parts[0])
        sale_val = parse_single_price(parts[1])
        if reg_val is not None:
            return {"regular": reg_val, "sale": sale_val}
        return None
    else:
        reg_val = parse_single_price(val_str)
        if reg_val is not None:
            return {"regular": reg_val, "sale": None}
        return None

def parse_price_cell(val: str):
    pp = parse_price_point(val)
    if pp:
        return pp.get("regular")
    return None


def parse_thc(val: str) -> str:
    if not val or val.strip() in ("", "-", "None"):
        return ""
    val = val.strip().replace("%", "")
    try:
        num = float(val)
        if num < 1:
            num = int(num * 100)
        else:
            num = int(num)
        return f"{num}%"
    except ValueError:
        return val


def detect_type(type_str: str) -> str:
    t = type_str.strip().upper()
    if t.startswith("I") or "INDICA" in t:
        return "indica"
    if t.startswith("S") or "SATIVA" in t:
        return "sativa"
    return "hybrid"


def normalize_sku(sku) -> str:
    """Normalize SKU: remove .0, strip whitespace."""
    return str(sku).replace(".0", "").strip()


# ── Flag Parsing ──

def parse_flower_flags(flag_rows: list[dict], store_code: str) -> dict:
    """
    Parse FlowersFlags_POS into a lookup dict.
    Returns: { sku: { "show": bool, "show5g": bool, "show14g": bool, "show28g": bool } }
    """
    flags = {}
    for row in flag_rows:
        store = row.get("Store", "").strip()
        if store != store_code:
            continue

        sku = normalize_sku(row.get("SKU", row.get("Key", "")))
        if not sku:
            continue

        show_main = row.get("Show", "SHOW").strip().upper()
        show5g = row.get("Show5G", "SHOW").strip().upper()
        show14g = row.get("Show14G", "SHOW").strip().upper()
        show28g = row.get("Show28G", "SHOW").strip().upper()
        status = row.get("Status", "").strip().upper()

        flags[sku] = {
            "show": show_main != "HIDDEN" and status != "SOLD OUT",
            "show5g": show5g != "HIDDEN",
            "show14g": show14g != "HIDDEN",
            "show28g": show28g != "HIDDEN",
        }

    return flags


def parse_item_flags(flag_rows: list[dict], store_code: str) -> dict:
    """
    Parse ItemsFlags_POS into a lookup dict.
    Returns: { sku: { "show": bool } }
    """
    flags = {}
    for row in flag_rows:
        store = row.get("Store", "").strip()
        if store != store_code:
            continue

        # Items can have multi-SKU keys like "800, 801, 802"
        key = normalize_sku(row.get("Key", ""))
        sku = normalize_sku(row.get("SKU", key))
        show = row.get("Show", "SHOW").strip().upper()

        # Use the Key field as identifier (matches ITEMS_LIVE SKU column)
        if key:
            flags[key] = {"show": show != "HIDDEN"}
        if sku and sku != key:
            flags[sku] = {"show": show != "HIDDEN"}

    return flags


# ── Builders ──

def build_flowers(product_rows: list[dict], flag_lookup: dict) -> list[dict]:
    """Build flowers.json from master data + store flags."""
    flowers = []
    skipped = 0

    for row in product_rows:
        sku = normalize_sku(row.get("SKU", ""))
        name = row.get("Strain", "").strip()
        if not sku or not name:
            continue

        tier = row.get("Tier", "").strip().upper()
        if tier not in ("EXOTIC", "PREMIUM", "AAA+", "AA", "BUDGET"):
            continue

        # Check store-specific flags
        flags = flag_lookup.get(sku)
        if not flags:
            skipped += 1
            continue

        # Item is shown if either main catalog show is True OR show28g is True (original TVMenu.html rule)
        show_main = flags.get("show", False)
        show_28g = flags.get("show28g", False)
        if not show_main and not show_28g:
            skipped += 1
            continue

        # Parse prices using parse_price_point
        price3g = parse_price_point(row.get("Price_3G", ""))
        price5g = parse_price_point(row.get("Price_5G", ""))
        price14g = parse_price_point(row.get("Price_14G", ""))
        price28g = parse_price_point(row.get("Price_28G", ""))

        # Apply weight visibility flags
        if not show_main and show_28g:
            # If hidden from main catalog but shown for 28g (e.g. shreds), nullify smaller weights
            price3g = None
            price5g = None
            price14g = None
        else:
            # Otherwise, apply standard weight visibility flags
            if not flags.get("show5g", True):
                price5g = None
            if not flags.get("show14g", True):
                price14g = None
            # Note: We do NOT nullify price28g if the item is generally visible (show_main=True)
            # and carries a valid 28g price in the spreadsheet. This ensures Sativas with 28g pricing
            # (like Daydream Diesel) populate their OZ prices and display on the OZ Card/Menu correctly!

        # Fallback to standard tier pricing if completely missing/null in sheet
        if price3g is None and price5g is None and price14g is None and price28g is None:
            if tier == "EXOTIC":
                price3g = {"regular": 40, "sale": None}
                price5g = {"regular": 60, "sale": None}
                price14g = {"regular": 140, "sale": None}
            elif tier == "PREMIUM":
                price3g = {"regular": 30, "sale": None}
                price5g = {"regular": 45, "sale": None}
                price14g = {"regular": 95, "sale": None}
            elif tier == "AAA+":
                price3g = {"regular": 20, "sale": None}
                price5g = {"regular": 30, "sale": None}
                price14g = {"regular": 70, "sale": None}
            elif tier == "AA":
                price5g = {"regular": 20, "sale": None}
                price14g = {"regular": 40, "sale": None}
            elif tier == "BUDGET":
                price3g = {"regular": 10, "sale": None}
                price28g = {"regular": 55, "sale": None}

        # Determine isSale from any price point having a sale price
        is_sale_detected = False
        for pp in (price3g, price5g, price14g, price28g):
            if pp and pp.get("sale") is not None:
                is_sale_detected = True
                break

        # Also check if name or Type contains "SALE"
        if not is_sale_detected:
            is_sale_detected = "SALE" in name.upper() or "SALE" in row.get("Type", "").upper()

        flower = {
            "sku": sku,
            "name": name,
            "slug": slugify(name),
            "tier": tier,
            "type": detect_type(row.get("Type", "hybrid")),
            "isHot": "HOT" in row.get("Type", "").upper() or "HOT" in name.upper(),
            "isSale": is_sale_detected,
            "thc": parse_thc(row.get("THC", "")),
            "price3g": price3g,
            "price5g": price5g,
            "price14g": price14g,
            "price28g": price28g,
            "image": row.get("ImageURL", "").strip(),
            "promoImage": row.get("PPromo", "").strip() or None,
        }
        flowers.append(flower)

    if skipped:
        print(f"  (Filtered out {skipped} hidden/sold-out flowers)")

    return flowers


def build_items(product_rows: list[dict], flag_lookup: dict) -> list[dict]:
    """Build items.json from master data + store flags."""
    items = []
    skipped = 0

    for row in product_rows:
        sku = row.get("SKU", "").strip()
        name = row.get("Name", "").strip()
        if not sku or not name:
            continue

        # Check store-specific flags (match on SKU key)
        sku_normalized = normalize_sku(sku)
        flags = flag_lookup.get(sku) or flag_lookup.get(sku_normalized)
        if not flags or not flags["show"]:
            skipped += 1
            continue

        price_raw = row.get("Price_EACH", "").strip()
        if price_raw:
            p = parse_price_cell(price_raw)
            price_str = f"${p}" if p else price_raw
        else:
            price_str = ""

        item = {
            "sku": sku_normalized,
            "name": name,
            "slug": slugify(name),
            "category": (row.get("Category", "") or "ADD ONS").strip().upper(),
            "type": row.get("Type", "").strip(),
            "thc": row.get("THC", "").strip(),
            "mg": row.get("MG", "").strip(),
            "price": price_str,
            "image": row.get("ImageURL", "").strip(),
            "promoImage": row.get("PPromoimageurl", row.get("PPromo", "")).strip() or None,
        }
        items.append(item)

    if skipped:
        print(f"  (Filtered out {skipped} hidden items)")

    return items


# ── Data Sources ──

def load_from_google_sheets():
    """Fetch all data from Google Sheets."""
    flower_products = fetch_sheet_csv(MENU_SHEET_ID, FLOWERS_SHEET)
    item_products = fetch_sheet_csv(MENU_SHEET_ID, ITEMS_SHEET)
    
    if POS_SHEET_ID:
        flower_flags = fetch_sheet_csv(POS_SHEET_ID, FLOWER_FLAGS_SHEET)
        item_flags = fetch_sheet_csv(POS_SHEET_ID, ITEM_FLAGS_SHEET)
    else:
        # Fall back to reading POS flags from local XLSX
        print("  Source (Flags): Local XLSX (POS_SHEET_ID not set)")
        import openpyxl
        menu_path = os.path.join(SCRIPT_DIR, "SETB Anti v2.3 6G MENU.xlsx")
        wb_menu = openpyxl.load_workbook(menu_path, data_only=True)
        flower_flags = []
        item_flags = []
        if FLOWER_FLAGS_SHEET in wb_menu.sheetnames:
            flower_flags = read_xlsx_sheet(wb_menu, FLOWER_FLAGS_SHEET)
        elif "WeightFlags" in wb_menu.sheetnames:
            flower_flags = read_xlsx_sheet(wb_menu, "WeightFlags")

        if ITEM_FLAGS_SHEET in wb_menu.sheetnames:
            item_flags = read_xlsx_sheet(wb_menu, ITEM_FLAGS_SHEET)
        elif "ItemsFlags" in wb_menu.sheetnames:
            item_flags = read_xlsx_sheet(wb_menu, "ItemsFlags")
            
    return flower_products, item_products, flower_flags, item_flags


def load_from_xlsx():
    """Fallback: read from local XLSX files."""
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: pip install openpyxl")
        sys.exit(1)

    # Menu sheet
    menu_path = os.path.join(SCRIPT_DIR, "SETB Anti v2.3 6G MENU.xlsx")
    if not os.path.exists(menu_path):
        print(f"  ERROR: {menu_path} not found")
        sys.exit(1)

    print(f"  Reading: {os.path.basename(menu_path)}")
    wb_menu = openpyxl.load_workbook(menu_path, data_only=True)
    flower_products = read_xlsx_sheet(wb_menu, FLOWERS_SHEET)
    item_products = read_xlsx_sheet(wb_menu, ITEMS_SHEET)

    # POS flags - check if in same workbook or separate
    flower_flags = []
    item_flags = []

    # Try reading flags from the menu workbook first
    if FLOWER_FLAGS_SHEET in wb_menu.sheetnames:
        flower_flags = read_xlsx_sheet(wb_menu, FLOWER_FLAGS_SHEET)
    elif "WeightFlags" in wb_menu.sheetnames:
        flower_flags = read_xlsx_sheet(wb_menu, "WeightFlags")

    if ITEM_FLAGS_SHEET in wb_menu.sheetnames:
        item_flags = read_xlsx_sheet(wb_menu, ITEM_FLAGS_SHEET)
    elif "ItemsFlags" in wb_menu.sheetnames:
        item_flags = read_xlsx_sheet(wb_menu, "ItemsFlags")

    # Check for separate POS XLSX
    for fname in os.listdir(SCRIPT_DIR):
        if "POS" in fname.upper() and fname.endswith(".xlsx"):
            pos_path = os.path.join(SCRIPT_DIR, fname)
            print(f"  Reading: {fname}")
            wb_pos = openpyxl.load_workbook(pos_path, data_only=True)
            if FLOWER_FLAGS_SHEET in wb_pos.sheetnames:
                flower_flags = read_xlsx_sheet(wb_pos, FLOWER_FLAGS_SHEET)
            if ITEM_FLAGS_SHEET in wb_pos.sheetnames:
                item_flags = read_xlsx_sheet(wb_pos, ITEM_FLAGS_SHEET)
            break

    return flower_products, item_products, flower_flags, item_flags


# ── Main ──

def main():
    print("=" * 60)
    print(f"  St Clair Cannabis - Product Sync (Store: {STORE_CODE})")
    print("=" * 60)
    print()

    # Load data
    if MENU_SHEET_ID:
        print("  Source: Google Sheets (live)")
        flower_products, item_products, flower_flags, item_flags = load_from_google_sheets()
    else:
        print("  Source: Local XLSX (set MENU_SHEET_ID for live sync)")
        flower_products, item_products, flower_flags, item_flags = load_from_xlsx()

    # Parse store-specific flags
    print(f"\n  Store flags for {STORE_CODE}:")
    flower_flag_lookup = parse_flower_flags(flower_flags, STORE_CODE)
    item_flag_lookup = parse_item_flags(item_flags, STORE_CODE)
    print(f"    Flower flags: {len(flower_flag_lookup)} SKUs configured")
    print(f"    Item flags:   {len(item_flag_lookup)} SKUs configured")

    # Build filtered JSONs
    print()
    flowers = build_flowers(flower_products, flower_flag_lookup)
    items = build_items(item_products, item_flag_lookup)

    # Stats
    print(f"\n  FLOWERS: {len(flowers)} active")
    tier_counts = {}
    for f in flowers:
        tier_counts[f["tier"]] = tier_counts.get(f["tier"], 0) + 1
    for tier in ("EXOTIC", "PREMIUM", "AAA+", "AA", "BUDGET"):
        print(f"    {tier}: {tier_counts.get(tier, 0)} strains")

    print(f"\n  ITEMS: {len(items)} active")
    cat_counts = {}
    for i in items:
        cat_counts[i["category"]] = cat_counts.get(i["category"], 0) + 1
    for cat, count in sorted(cat_counts.items()):
        print(f"    {cat}: {count}")

    # Write JSONs
    with open(FLOWERS_OUT, "w", encoding="utf-8") as f:
        json.dump(flowers, f, indent=2, ensure_ascii=False)
    print(f"\n  [OK] flowers.json ({len(flowers)} products)")

    with open(ITEMS_OUT, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)
    print(f"  [OK] items.json ({len(items)} products)")

    # Deploy
    print()
    if "--deploy" in sys.argv:
        do_deploy = True
    else:
        try:
            ans = input("  Deploy to Vercel now? (y/n): ").strip().lower()
            do_deploy = ans in ("y", "yes")
        except EOFError:
            do_deploy = False

    if do_deploy:
        print("\n  Building + deploying...")
        result = subprocess.run(
            ["powershell", "-Command",
             "Set-ExecutionPolicy -Scope Process -ExecutionPolicy Bypass; "
             "npm run build 2>&1; vercel --prod --yes 2>&1"],
            cwd=SCRIPT_DIR,
            capture_output=True, text=True
        )
        if result.returncode == 0:
            print("  [OK] Live at https://stclaircannabis.com")
        else:
            print(f"  Output:\n{result.stdout[-500:]}")
    else:
        print("  To deploy: npm run build && vercel --prod --yes")

    print("\n  Done!")


if __name__ == "__main__":
    main()
